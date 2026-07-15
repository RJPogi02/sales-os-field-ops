from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\HKJ-ITLP-003\Downloads\Batching Plants CRM System - Working Copy.xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "src" / "data" / "leads.json"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def key(value):
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def first(row, headers, candidates):
    normalized = {key(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate)
        if index is not None and index < len(row):
            value = clean(row[index])
            if value:
                return value
    for header, index in normalized.items():
        if any(candidate in header for candidate in candidates):
            value = clean(row[index])
            if value:
                return value
    return ""


def header_row_index(rows):
    signals = {
        "company", "company name", "batching plant", "location", "address",
        "phone", "contact number", "email", "status", "notes", "region"
    }
    scored = []
    for index, row in enumerate(rows[:12]):
        score = sum(1 for value in row if key(value) in signals)
        score += sum(1 for value in row if any(signal in key(value) for signal in signals))
        scored.append((score, index))
    return max(scored, default=(0, 0))[1]


def extract():
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    preferred = [name for name in workbook.sheetnames if key(name) in {"north", "ncr", "south", "leads normalized"}]
    sheet_names = preferred or workbook.sheetnames
    leads = []

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        header_index = header_row_index(rows)
        headers = rows[header_index]
        for source_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            company = first(row, headers, ["company name", "company", "batching plant", "plant name", "name"])
            phone = first(row, headers, ["phone", "contact number", "mobile", "telephone", "tel"])
            email = first(row, headers, ["email", "email address", "e mail"])
            location = first(row, headers, ["location", "address", "city municipality", "city", "municipality"])
            if not company or not any([phone, email, location]):
                continue

            raw_score = first(row, headers, ["lead score", "score", "priority"])
            score_digits = re.findall(r"\d+", raw_score)
            numeric_score = int(score_digits[0]) if score_digits else 0
            priority = "High" if numeric_score >= 70 else "Medium" if numeric_score >= 35 else "Low"
            if key(raw_score) in {"high", "medium", "low"}:
                priority = raw_score.title()

            email_value = first(row, headers, ["email", "email address", "e mail"])
            facebook_value = first(row, headers, ["facebook", "facebook link", "fb"])
            if email_value and "@" not in email_value:
                if not facebook_value and "facebook" in email_value.lower():
                    facebook_value = email_value
                email_value = ""

            leads.append({
                "id": f"{key(sheet_name).replace(' ', '-')}-{source_row}",
                "company": company,
                "region": first(row, headers, ["region", "area"]) or sheet_name.upper(),
                "location": location,
                "phone": phone,
                "email": email_value,
                "contactPerson": first(row, headers, ["contact person", "contact", "procurement", "decision maker"]),
                "facebook": facebook_value,
                "maps": first(row, headers, ["google maps", "maps", "maps link"]),
                "status": first(row, headers, ["status", "lead status"]) or "New Lead",
                "lastResult": first(row, headers, ["call result", "last result", "result"]) or "Not called",
                "lastContacted": first(row, headers, ["last contacted", "last contact", "date contacted"]),
                "nextFollowUp": first(row, headers, ["next follow up", "follow up date", "follow up"]),
                "notes": first(row, headers, ["notes", "remarks", "call notes"]),
                "priority": priority,
                "leadScore": numeric_score,
                "selected": False,
                "quoteReady": False,
                "profileSent": False,
                "answered": False,
                "checklist": [False, False, False, False, False, False, False],
            })

    seen = set()
    unique = []
    for lead in leads:
        signature = (key(lead["company"]), re.sub(r"\D", "", lead["phone"]))
        if signature in seen:
            continue
        seen.add(signature)
        unique.append(lead)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(unique, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"source": str(SOURCE), "sheets": sheet_names, "lead_count": len(unique), "output": str(OUTPUT)}))


if __name__ == "__main__":
    extract()
